import openpyxl
import os

def unmerge_and_save_excel(filepath):
    """
    拆分Excel文件中的合并单元格，并保存到新文件，然后删除原文件。

    参数:
    filepath (str): Excel文件的路径。
    """
    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # 记录需要拆分的合并单元格和对应的值
        merged_cells_data = []
        for merged_cells in sheet.merged_cells.ranges:
            top_left_cell = merged_cells.min_row, merged_cells.min_col
            value = sheet.cell(row=top_left_cell[0], column=top_left_cell[1]).value
            merged_cells_data.append((merged_cells, value))

        # 拆分合并单元格
        for merged_cells, value in merged_cells_data:
            sheet.unmerge_cells(range_string=str(merged_cells))
            for row in range(merged_cells.min_row, merged_cells.max_row + 1):
                for col in range(merged_cells.min_col, merged_cells.max_col + 1):
                    sheet.cell(row=row, column=col, value=value)
        # 删除原始文件
        os.remove(filepath)
        # 保存拆分后的Excel文件
        new_filepath = os.path.splitext(filepath)[0] + ".xlsx"
        workbook.save(new_filepath)

        
        print(f"成功拆分并保存到 {new_filepath}，并删除了原始文件。")
    except FileNotFoundError:
        print(f"错误：文件 {filepath} 未找到。")
    except Exception as e:
        print(f"发生错误：{e}")

# 示例用法：
filepath = "Try-Markdown-RAG/excels/个人简历.xlsx"  # 替换为你的Excel文件路径
unmerge_and_save_excel(filepath)