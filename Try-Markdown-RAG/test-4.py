import pandas as pd
import jieba
import re
import traceback
import torch
from typing import List, Optional, Dict
from langchain_core.documents import Document
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from transformers import AutoTokenizer, AutoModel
import time  # 新增导入

class LocalLLM:
    """本地部署的中文LLM模型"""
    def __init__(self, model_path: str = "/opt/share/models/thu-coai/ShieldLM-6B-chatglm3/snapshots/7b1f7843a3803fdd35b7a2013d5640eca3eca2ee"):
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.tokenizer = AutoTokenizer.from_pretrained(
            model_path, 
            trust_remote_code=True
        )
        self.model = AutoModel.from_pretrained(
            model_path,
            trust_remote_code=True,
            device_map="auto",
            torch_dtype=torch.float16
        ).eval()
    
    def generate(self, prompt: str, max_length: int = 1000) -> str:
        """生成文本"""
        response, _ = self.model.chat(
            self.tokenizer,
            prompt,
            history=[],
            max_length=max_length,
            temperature=0.2
        )
        return response
class ExcelAnalyzer:
    """增强版Excel结构分析器"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.analysis = {
            "file_info": {},
            "sheets": {}
        }
        
    def _detect_merged_cells(self, sheet):
        """检测合并单元格"""
        merged_ranges = []
        for merged_range in sheet.merged_cells.ranges:
            merged_ranges.append({
                "range": str(merged_range),
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col
            })
        return merged_ranges
    
    def analyze(self) -> Dict:
        """执行完整分析"""
        try:
            # 使用openpyxl获取更详细的信息
            wb = load_workbook(self.file_path)
            
            self.analysis["file_info"] = {
                "file_type": "Excel",
                "sheets_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames
            }
            
            # 分析每个工作表
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                df = pd.read_excel(self.file_path, sheet_name=sheet_name, nrows=5)
                
                # 获取列元数据
                columns_meta = []
                for idx, col in enumerate(df.columns, 1):
                    sample_values = df[col].head(3).fillna("").tolist()
                    dtype = str(df[col].dtype)
                    columns_meta.append({
                        "position": get_column_letter(idx+1),
                        "name": str(col),
                        "data_type": dtype,
                        "sample_values": sample_values
                    })
                
                # 存储分析结果
                self.analysis["sheets"][sheet_name] = {
                    "dimension": f"{sheet.max_row}行×{sheet.max_column}列",
                    "merged_cells": self._detect_merged_cells(sheet),
                    "columns": columns_meta,
                    "header_candidates": self._find_header_candidates(df),
                    "data_regions": self._detect_data_regions(sheet)
                }
            
            return self.analysis
        
        except Exception as e:
            print(f"文件分析失败: {str(e)}")
            return {}
    
    def _find_header_candidates(self, df: pd.DataFrame) -> List[int]:
        """智能识别可能的表头行"""
        candidates = []
        for i in range(min(3, len(df))):
            # 检查是否包含常见表头特征
            row = df.iloc[i].astype(str)
            if any(re.match(r"列\d+|column|项目|名称", x, re.IGNORECASE) for x in row):
                candidates.append(i+1)  # 返回Excel行号
        return candidates or [1]
    
    def _detect_data_regions(self, sheet) -> List[Dict]:
        """检测数据区域"""
        regions = []
        # 简单实现：检测连续非空区域
        max_row = min(sheet.max_row, 1000)
        max_col = min(sheet.max_column, 50)
        
        current_region = None
        for row in range(1, max_row+1):
            has_data = any(sheet.cell(row=row, column=col).value for col in range(1, max_col+1))
            
            if has_data:
                if not current_region:
                    current_region = {"start_row": row, "end_row": row}
                else:
                    current_region["end_row"] = row
            else:
                if current_region:
                    regions.append(current_region)
                    current_region = None
        return regions
    
class CodeGenerator:
    """基于本地LLM的代码生成器"""
    
    def __init__(self, model_path: str = "THUDM/chatglm3-6b"):
        self.llm = LocalLLM(model_path)
        
    def generate_code(self, query: str, analysis: Dict) -> Optional[str]:
        """生成检索代码"""
        prompt = self._build_prompt(query, analysis)
        try:
            response = self.llm.generate(prompt)
            return self._extract_code(response)
        except Exception as e:
            print(f"代码生成失败: {str(e)}")
            return None
    
    def _build_prompt(self, query: str, analysis: Dict) -> str:
        """构建LLM提示词"""
        analysis_str = self._format_analysis(analysis)
        
        return f"""你是一个数据分析专家，请根据用户查询和Excel结构分析结果，生成正确的Python代码来提取所需信息。

                # 用户需求：
                {query}

                # Excel结构分析：
                {analysis_str}

                # 代码要求：
                1. 使用pandas处理Excel文件
                2. 优先使用识别到的表头信息
                3. 处理可能的合并单元格和复杂结构
                4. 结果保存到result变量
                5. 最终输出为自然语言格式

                # 响应格式：
                请按以下格式生成代码：
                ```python
                import pandas as pd

                excel_file = pd.ExcelFile('{{file_path}}')
                result = []

                # 你的处理逻辑...

                result = '\\n'.join(result)
                """

    def _format_analysis(self, analysis: Dict) -> str:
        """格式化分析结果"""
        output = []
        for sheet_name, info in analysis["sheets"].items():
            output.append(f"工作表【{sheet_name}】")
            output.append(f"- 维度: {info['dimension']}")
            output.append(f"- 可能表头行: {info['header_candidates']}")
            output.append("列信息:")
            for col in info["columns"]:
                samples = ", ".join(map(str, col["sample_values"][:3]))
                output.append(f"  {col['position']}列 [{col['data_type']}] {col['name']} (示例值: {samples})")
        return "\n".join(output)
    
    def _extract_code(self, text: str) -> str:
        """从响应中提取代码块"""
        code_blocks = re.findall(r'```python(.*?)```', text, re.DOTALL)
        return code_blocks[0].strip() if code_blocks else ""

class ExcelRetriever:
    """改进版Excel检索器"""

    def __init__(self, model_path: str = "THUDM/chatglm3-6b"):
        self.code_gen = CodeGenerator(model_path)
        self.safety_check = CodeSafetyChecker()
    def retrieve(self, query: str, file_path: str) -> Optional[List[Document]]:
        """执行完整检索流程"""
        try:
            print(f"\n开始分析Excel文件: {file_path}")
            # 步骤1：分析Excel结构
            analyzer = ExcelAnalyzer(file_path)
            analysis = analyzer.analyze()
            if not analysis["sheets"]:
                print("Excel分析失败：未找到有效的工作表")
                return None
            print("Excel结构分析完成")
            
            # 步骤2：生成检索代码
            print("\n正在生成检索代码...")
            code = self.code_gen.generate_code(query, analysis)
            if not code:
                print("代码生成失败")
                return None
            print("\n生成的检索代码：")
            print("```python")
            print(code)
            print("```")
                
            # 步骤3：安全检查和执行
            print("\n正在进行代码安全检查...")
            if not self.safety_check.validate(code):
                print("代码安全检查未通过")
                return None
            print("代码安全检查通过")
                
            print("\n开始执行检索代码...")
            return self._execute_code(code, file_path)
            
        except Exception as e:
            print(f"检索过程出错: {str(e)}")
            traceback.print_exc()
            return None
    def _execute_code(self, code: str, file_path: str) -> Optional[List[Document]]:
        """在隔离环境中执行代码"""
        local_vars = {
            "pd": pd,
            "__file__": file_path,
            "result": None
        }
        
        try:
            # 替换文件路径占位符
            final_code = code.replace("{{file_path}}", repr(file_path))
            
            # 限制可用的模块和函数
            allowed_builtins = {
                'range': range,
                'len': len,
                'str': str,
                'int': int,
                'float': float
            }
            
            exec_globals = {
                "__builtins__": allowed_builtins,
                "pd": pd
            }
            
            exec(final_code, exec_globals, local_vars)
            
            if local_vars["result"]:
                return [Document(page_content=local_vars["result"])]
                
        except Exception as e:
            print(f"代码执行出错: {str(e.__class__.__name__)} - {str(e)}")
            traceback.print_exc()
        
        return None

class CodeSafetyChecker:
    """代码安全检查器"""

    def __init__(self):
        self.allowed_modules = {"pandas": ["pd"]}
        self.banned_keywords = {
            "os.", "sys.", "subprocess", "shutil", 
            "open(", "__import__", "eval(", "exec("
        }

    def validate(self, code: str) -> bool:
        """检查代码安全性"""
        # 检查危险关键词
        for keyword in self.banned_keywords:
            if keyword in code:
                print(f"检测到危险操作: {keyword}")
                return False
                
        # 检查允许的模块
        imports = re.findall(r'^\s*import (\w+)', code, re.M)
        for imp in imports:
            if imp not in self.allowed_modules:
                print(f"禁止导入模块: {imp}")
                return False
                
        return True
    def safe_retrieve(self, query: str, file_path: str, model_path: str) -> str:
        """安全检索入口"""
        try:
            retriever = ExcelRetriever(model_path)
            results = retriever.retrieve(query, file_path)
            return results[0].page_content if results else "未找到相关信息"
        except Exception as e:
            print(f"检索失败: {str(e)}")
            return "检索过程中发生错误"

if __name__ == "__main__":
    # 测试配置
    TEST_FILE = "Try-Markdown-RAG/excels/个人简历.xlsx"
    # 测试用例
    test_query = "请提取张三的实践经历信息"

    print("正在初始化模型...")
    start_time = time.time()
    safety_checker = CodeSafetyChecker()
    result = safety_checker.safe_retrieve(
        query=test_query,
        file_path=TEST_FILE,
        model_path="/opt/share/models/thu-coai/ShieldLM-6B-chatglm3/snapshots/7b1f7843a3803fdd35b7a2013d5640eca3eca2ee"
    )

    print(f"\n检索耗时: {time.time()-start_time:.2f}秒")
    print("\n最终结果：")
    print(result)