import os
import pandas as pd
from langchain.docstore.document import Document
import openpyxl

def unmerge_and_save_excel(filepath):
    """
    拆分Excel文件中的合并单元格，并保存到新文件，然后删除原文件。

    参数:
    filepath (str): Excel文件的路径。
    """
    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # 记录需要拆分的合并单元格和对应的值
        merged_cells_data = []
        for merged_cells in sheet.merged_cells.ranges:
            top_left_cell = merged_cells.min_row, merged_cells.min_col
            value = sheet.cell(row=top_left_cell[0], column=top_left_cell[1]).value
            merged_cells_data.append((merged_cells, value))

        # 拆分合并单元格
        for merged_cells, value in merged_cells_data:
            sheet.unmerge_cells(range_string=str(merged_cells))
            for row in range(merged_cells.min_row, merged_cells.max_row + 1):
                for col in range(merged_cells.min_col, merged_cells.max_col + 1):
                    sheet.cell(row=row, column=col, value=value)
        # 删除原始文件
        os.remove(filepath)
        # 保存拆分后的Excel文件
        new_filepath = os.path.splitext(filepath)[0] + ".xlsx"
        workbook.save(new_filepath)

        print(f"成功拆分并保存到 {new_filepath}，并删除了原始文件。")
    except FileNotFoundError:
        print(f"错误：文件 {filepath} 未找到。")
    except Exception as e:
        print(f"发生错误：{e}")


class ExcelParser:
    def __init__(self, chunk_size=1000):
        self.chunk_size = chunk_size

    def parse_excel_to_documents(self, file_path):
        """解析 Excel 文件并生成文档分块"""
        print(f"开始解析 Excel 文件: {file_path}")
        
        # 读取 Excel 文件
        df = pd.read_excel(file_path, sheet_name=None)
        
        documents = []
        
        for sheet_name, sheet_data in df.items():
            print(f"解析工作表: {sheet_name}")
            
            current_chunk = ""
            current_length = 0
            
            for index, row in sheet_data.iterrows():
                row_content = self._generate_chunk_from_row(sheet_data.columns, row)
                row_length = len(row_content)
                
                # 计算添加后的总长度（包含换行符）
                if current_chunk:
                    potential_length = current_length + 1 + row_length  # 换行符占1个字符
                else:
                    potential_length = row_length
                
                if potential_length <= self.chunk_size:
                    # 追加到当前块
                    if current_chunk:
                        current_chunk += "\n" + row_content
                        current_length += 1 + row_length
                    else:
                        current_chunk = row_content
                        current_length = row_length
                else:
                    # 先保存已有块
                    if current_chunk:
                        documents.append(Document(page_content=current_chunk))
                        current_chunk = ""
                        current_length = 0
                    
                    # 处理当前行
                    if row_length > self.chunk_size:
                        print(f"行内容过长（{row_length}字符），将进行分块处理")
                        split_chunks = self._split_long_content(row_content)
                        documents.extend([Document(page_content=chunk) for chunk in split_chunks])
                    else:
                        current_chunk = row_content
                        current_length = row_length
            
            # 处理最后一个块
            if current_chunk:
                documents.append(Document(page_content=current_chunk))
        
        print(f"共生成 {len(documents)} 个文档分块")
        return documents

    def _generate_chunk_from_row(self, columns, row):
        """生成单行内容字符串"""
        chunk = ""
        for col, value in row.items():
            chunk += f"{col}: {value}  "
        return chunk.strip()

    def _split_long_content(self, content):
        """拆分超长行内容"""
        chunks = []
        while len(content) > self.chunk_size:
            # 找最后一个空格进行拆分
            split_pos = content.rfind(' ', 0, self.chunk_size)
            if split_pos == -1:
                split_pos = self.chunk_size
            chunks.append(content[:split_pos])
            content = content[split_pos:].lstrip()
        if content:
            chunks.append(content)
        return chunks